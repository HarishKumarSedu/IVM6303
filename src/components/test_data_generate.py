import re 
import json
import openpyxl
import yaml
import pandas as pd 
import os
from src import log 
from src.utils.common import read_yaml,create_directories
from src.constants import CONFIG_FILE_PATH

class TestDataGenerate:
    def __init__(self, config_file_path=CONFIG_FILE_PATH) -> None:
        self.config = read_yaml(config_file_path)
        self.data_ingestion = self.config.data_ingestion 
        self.test_data_generate = self.config.test_data_generate 
        #create the test data directory 
        if not os.path.exists(self.test_data_generate.root_dir):
            os.makedires(self.test_data_generate.root_dir)
            log.info('test_data_generate directory created')

    def generate_test_data(self):
        # determine the root data file from the the root directory and the local file name 
        data_root_file = os.path.join(self.data_ingestion.root_dir, self.data_ingestion.local_filename)
        sheetnames =  openpyxl.load_workbook(data_root_file).sheetnames
        config = {}
        for sheet in sheetnames:
            data = pd.read_excel(data_root_file,sheet_name=sheet)
            dataset = {}
            sheet = re.sub('[!-\/:-@[-`{-~\s]','_', sheet) 
            tests = []
            print(data.columns[1:])
            for column in data.columns[1:]:
                tests.append(re.sub('[!-\/:-@[-`{-~\s]','_',data[column][2]))
                dataset.update(
                    {
                        tests[-1]: {
                            "TestNumber" :data[column][1],
                            "TestStatus" :data[column][0],
                            "TestData" :data[column][3],
                            "Limits" : {
                                "Min" : data[column][4],
                                "Min Value/%" : data[column][5],
                                "Min Value/%" : data[column][5],
                                "Value" : data[column][6],
                                "Max" : data[column][7],
                                "Max Value/%" : data[column][8],
                                },
                        }
                    }
                )
            config.update({
                sheet:tests
            })
            with open(os.path.join(self.test_data_generate.root_dir, self.test_data_generate.test_data_filename),'w') as file:
                json.dump(dataset, file)
        with open( os.path.join(os.path.join(self.test_data_generate.root_dir, self.test_data_generate.test_cofig_filename)),'w') as file:
            yaml.dump({"TestBench":config}, file)